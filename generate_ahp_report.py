from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


out = Path(r"C:\Users\youhs\Documents\實習專案")
html_path = out / "v3-tech-radar-ahp-scoring-report.html"
xlsx_path = out / "v3-tech-radar-ahp-scoring-summary.xlsx"

criteria = [
    {"id": "maturity", "zh": "技術成熟度", "weight": 0.35, "s4_weight": 0.20, "basis": "NASA TRL、Thoughtworks Radar", "meaning": "技術是否已從概念、preview、GA 走到企業可採用階段。"},
    {"id": "aws_fit", "zh": "AWS / 公司情境適配度", "weight": 0.25, "s4_weight": 0.20, "basis": "AWS Well-Architected、TOE Framework", "meaning": "是否直接支撐 AWS workload、公司雲端治理與國泰情境。"},
    {"id": "case_evidence", "zh": "企業案例證據", "weight": 0.15, "s4_weight": 0.20, "basis": "Rogers Diffusion、Thoughtworks Radar", "meaning": "是否已有企業、金融、保險或受監管產業案例可觀察。"},
    {"id": "effort", "zh": "導入難度", "weight": 0.15, "s4_weight": 0.20, "basis": "Rogers complexity、TOE、TIME、MCDM", "meaning": "PoC 或導入所需權限、跨服務整合、組織協調與成本。"},
    {"id": "risk", "zh": "導入風險", "weight": 0.10, "s4_weight": 0.20, "basis": "NIST AI RMF、ISO 31000、AWS Well-Architected", "meaning": "資安、合規、營運、資料與 AI 模型風險。"},
]

rubric_rows = [
    ["技術成熟度", "概念或 preview", "早期服務 / 文件不足", "已 GA 但案例有限", "已有穩定企業採用", "金融或大型企業成熟採用"],
    ["AWS / 公司情境適配度", "與 AWS 雷達弱相關", "間接相關", "可放入 AWS 架構但非核心", "明確強化 AWS workload", "直接解決公司 AWS 架構痛點"],
    ["企業案例證據", "無企業案例", "有一般技術文章", "有企業案例", "有金融 / 保險相近案例", "有大型金融 / 受監管產業案例"],
    ["導入難度", "很容易試，幾乎無新權限", "小型 PoC 可做", "需跨服務整合", "需資安 / 架構 / 權限協調", "高成本或高組織阻力"],
    ["導入風險", "低風險", "可控風險", "需一般治理", "高資安 / 合規 / 營運風險", "不建議近期導入"],
]

sources = [
    ["NASA Technology Readiness Levels", "官方成熟度量表", "maturity", "用 TRL 成熟度概念支撐 1-5 成熟度分層。", "https://www.nasa.gov/directorates/somd/space-communications-navigation-program/technology-readiness-levels/"],
    ["Thoughtworks Technology Radar", "企業技術雷達方法", "maturity, case_evidence", "用 Adopt / Trial / Assess / Hold 支撐技術雷達採用狀態。", "https://www.thoughtworks.com/radar"],
    ["AWS Well-Architected Framework", "AWS 官方架構評估", "aws_fit, effort, risk", "六大支柱支撐 AWS 適配、成本、資安、可靠性與營運風險。", "https://docs.aws.amazon.com/wellarchitected/latest/framework/the-pillars-of-the-framework.html"],
    ["NIST AI Risk Management Framework", "官方 AI 風險治理", "risk", "Govern / Map / Measure / Manage 支撐 AI 導入風險治理。", "https://www.nist.gov/itl/ai-risk-management-framework"],
    ["ISO 31000 Risk Management", "國際風險管理標準", "risk", "支撐風險識別、分析、評估、處置與監控。", "https://www.iso.org/standard/65694.html"],
    ["TOE Framework", "企業技術採用理論", "aws_fit, effort", "Technology-Organization-Environment 說明採用受技術、組織與環境脈絡影響。", "https://academic-publishing.org/index.php/ejise/article/download/389/352/385"],
    ["Rogers Diffusion of Innovations", "創新擴散理論", "case_evidence, effort", "observability、trialability、complexity 支撐案例證據與導入難度。", "https://open.ncl.ac.uk/theories/8/diffusion-of-innovations/"],
    ["MC2 cloud MCDM", "雲端多準則決策文獻", "all", "雲端採用不應只看成本，需同時評估 benefits、opportunities、risks。", "https://arxiv.org/abs/1112.1851"],
    ["SAP LeanIX TIME model", "企業 portfolio rationalization", "aws_fit, effort, risk", "用 business/technical fit、成本、風險支撐 IT 投資優先級。", "https://help.sap.com/docs/leanix/ea/time"],
]

weights = [c["weight"] for c in criteria]
names = [c["zh"] for c in criteria]
matrix = [[weights[i] / weights[j] for j in range(len(weights))] for i in range(len(weights))]

stats = [
    ["外部來源總數", 9, "官方框架、學術來源、企業方法論合計。"],
    ["官方 / 標準來源", 4, "NASA TRL、AWS Well-Architected、NIST AI RMF、ISO 31000。"],
    ["學術 / 理論來源", 3, "TOE、Rogers Diffusion、MC2 cloud MCDM。"],
    ["企業方法來源", 2, "Thoughtworks Technology Radar、SAP LeanIX TIME。"],
    ["AHP-like 準則數", 5, "maturity、aws_fit、case_evidence、effort、risk。"],
    ["S3 最大權重", "0.35", "技術成熟度。"],
    ["S3 第二權重", "0.25", "AWS / 公司情境適配度。"],
    ["S3 反向計分指標", 2, "effort 與 risk 使用 6 - score。"],
    ["一致性比率 CR", "0.000", "此矩陣由權重向量反推，完全一致；後續可用 mentor pairwise survey 校準。"],
]

with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
    pd.DataFrame([[c["id"], c["zh"], c["weight"], c["s4_weight"], c["basis"], c["meaning"]] for c in criteria], columns=["criteria_id", "指標", "S3 AHP-like 權重", "S4 複核權重", "參照來源", "定義"]).to_excel(writer, index=False, sheet_name="weights")
    pd.DataFrame([[names[i]] + [round(v, 3) for v in row] for i, row in enumerate(matrix)], columns=["criteria"] + names).to_excel(writer, index=False, sheet_name="pairwise_matrix")
    pd.DataFrame(rubric_rows, columns=["指標", "1分", "2分", "3分", "4分", "5分"]).to_excel(writer, index=False, sheet_name="rubric_scale")
    pd.DataFrame(sources, columns=["來源", "類型", "支撐指標", "用途", "URL"]).to_excel(writer, index=False, sheet_name="sources")
    pd.DataFrame(stats, columns=["統計項目", "數值", "說明"]).to_excel(writer, index=False, sheet_name="summary_stats")

wb = load_workbook(xlsx_path)
for ws in wb.worksheets:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"), top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB"))
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = min(48, max(12, max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, min(ws.max_row, 30) + 1)) + 2))
wb.save(xlsx_path)


def head(cols):
    return "<tr>" + "".join(f"<th>{c}</th>" for c in cols) + "</tr>"


def rows(data):
    return "\n".join("<tr>" + "".join(f"<td>{cell}</td>" for cell in row) + "</tr>" for row in data)


weights_rows = [[c["zh"], c["id"], f'{c["weight"]:.2f}', f'{c["s4_weight"]:.2f}', c["basis"], c["meaning"]] for c in criteria]
matrix_rows = [[names[i]] + [f"{matrix[i][j]:.3f}" for j in range(len(criteria))] for i in range(len(criteria))]
sources_rows = [[s[0], s[1], s[2], s[3], f'<a href="{s[4]}">{s[4]}</a>'] for s in sources]

css = """
@page { size: A4; margin: 14mm 12mm; }
body { font-family: 'Microsoft JhengHei', 'Noto Sans TC', Arial, sans-serif; color:#111827; line-height:1.55; }
h1 { text-align:center; font-size:26px; margin:0 0 4px; }
.subtitle { text-align:center; color:#4b5563; margin-bottom:22px; }
h2 { font-size:19px; border-left:5px solid #2563eb; padding-left:10px; margin-top:22px; }
p, li { font-size:12.5px; }
table { width:100%; border-collapse:collapse; margin:10px 0 16px; page-break-inside:auto; }
th, td { border:1px solid #d1d5db; padding:6px 7px; font-size:11px; vertical-align:top; }
th { background:#1f2937; color:white; text-align:center; }
tr:nth-child(even) td { background:#f9fafb; }
.formula { background:#f3f4f6; padding:9px 11px; border-radius:6px; font-family:Consolas, monospace; font-size:11px; }
.note { border:1px solid #bfdbfe; background:#eff6ff; padding:10px 12px; border-radius:6px; font-size:12px; }
.page-break { page-break-before:always; }
a { color:#1d4ed8; text-decoration:none; }
"""

html = f"""<!doctype html>
<html lang="zh-Hant">
<head><meta charset="utf-8"><title>v3 技術雷達 AHP-like 評分量度分析報告</title><style>{css}</style></head>
<body>
<h1>v3 技術雷達 AHP-like 評分量度分析報告</h1>
<div class="subtitle">公司帳戶落地版｜rubric fallback 評分方法與文獻佐證｜2026-07-14</div>

<h2>1. 報告目的</h2>
<p>本報告重整 v3 技術雷達在 LLM 無法使用、API key 尚未設定、或報價超過上限時的評分方法。新版採用 <b>AHP-like multi-criteria decision model</b>：先定義決策目標，再建立五個準則，最後以文獻與企業方法論校準權重與 1-5 分量尺。</p>
<div class="note">定位：這不是宣稱已完成正式 AHP 專家問卷，而是以 AHP 的層級化、多準則、權重透明精神，建立可解釋的工程落地版評分模型。若後續 mentor 願意參與，可再用 pairwise comparison 問卷重新校準權重。</div>

<h2>2. AHP-like 決策階層</h2>
<table>{head(['層級','內容','說明'])}
{rows([
['目標層', '選出每日 AWS 技術雷達 Top 3', '在成熟度、AWS 適配、案例證據、導入難度與風險之間取得平衡。'],
['準則層', 'maturity / aws_fit / case_evidence / effort / risk', '五項準則分別對應技術成熟度、雲端架構適配、外部案例、導入摩擦與治理風險。'],
['方案層', '每日候選技術或 AWS 新聞項目', 'S1/S2 篩選後進入 S3/S4 評分，最後由平均分數排序。']
])}</table>

<h2>3. 權重設計與文獻參照</h2>
<p>S3 evaluator 偏重技術成熟度與 AWS 適配，因為第一輪評估要快速判斷「是否值得進一步 PoC」。S4 validator 採五項平均權重，作為獨立複核，避免第一輪權重過度偏向成熟技術而壓低創新項目。</p>
<table>{head(['指標','程式欄位','S3 權重','S4 權重','主要參照來源','定義'])}{rows(weights_rows)}</table>
<div class="formula">S3 score = maturity*0.35 + aws_fit*0.25 + case_evidence*0.15 + (6-effort)*0.15 + (6-risk)*0.10<br>S4 score = maturity*0.20 + aws_fit*0.20 + case_evidence*0.20 + (6-effort)*0.20 + (6-risk)*0.20<br>Final score = (S3 score + S4 score) / 2</div>

<h2>4. AHP-like Pairwise Matrix</h2>
<p>下表以 S3 權重向量反推 pairwise comparison matrix。因為矩陣由同一組權重生成，所以完全一致，λmax = 5，CI = 0.000，CR = 0.000。實務上，這可作為初版 baseline；後續可以請 mentor 直接調整兩兩比較值。</p>
<table>{head(['準則'] + names)}{rows(matrix_rows)}</table>

<h2>5. 1-5 分量尺</h2>
<table>{head(['指標','1 分','2 分','3 分','4 分','5 分'])}{rows(rubric_rows)}</table>

<div class="page-break"></div>
<h2>6. 外部來源與佐證對應</h2>
<p>本模型參照官方框架、標準、學術理論與企業方法論。正式說明時建議用「參考 / 佐證 / 映射」措辭，而不是說模型直接由單一論文推導。</p>
<table>{head(['來源','類型','支撐指標','用途','連結'])}{rows(sources_rows)}</table>

<h2>7. 統計彙整表</h2>
<table>{head(['統計項目','數值','說明'])}{rows(stats)}</table>

<h2>8. 結論與使用限制</h2>
<p>新版評分模型可以更清楚說明：為什麼成熟度、AWS 適配與企業案例會影響 Top 3 排序，也能解釋為什麼導入難度與風險採反向計分。它適合早期技術雷達排序、fallback scoring、mentor 討論與 PoC 候選篩選。</p>
<p>限制是：此模型仍屬初版 engineering rubric，不應取代正式資安、法遵、採購或架構審查。進入正式導入前，建議由 mentor、架構師、資安與業務單位共同校準權重。</p>
</body></html>"""

html_path.write_text(html, encoding="utf-8")
print(html_path)
print(xlsx_path)
